import openpyxl
import tkinter as tk

# Exportar os dados em planilhas
def xl_export(listbox, nome_arquivo):

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ["Razão Social", "Telefone", "Email", "Bairro/Município", "Representante"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col).value = header

    for i, item in enumerate(listbox.get(0, tk.END), start=2):
        colunas = item.split(" | ")

        for j, coluna in enumerate(colunas, start=1):
            sheet.cell(row=i, column=j).value = coluna

    workbook.save(nome_arquivo)

def xl_import(array_dados_cadastrados, planilha):

    workbook = openpyxl.load_workbook(planilha)
    sheet = workbook.active

    # Ler dados cadastrados
    array_dados_cadastrados.clear()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        array_dados_cadastrados.append(list(row))